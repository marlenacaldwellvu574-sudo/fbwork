import os
import logging
from telegram import Update
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    ContextTypes, ConversationHandler, filters
)
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
import time
import json

BOT_TOKEN = os.getenv("BOT_TOKEN")
_admin_id  = os.getenv("ADMIN_ID")

if not BOT_TOKEN:
    raise SystemExit("ERROR: BOT_TOKEN environment variable is not set.")
if not _admin_id:
    raise SystemExit("ERROR: ADMIN_ID environment variable is not set. Get your numeric ID from @userinfobot on Telegram.")
try:
    ADMIN_ID = int(_admin_id)
except ValueError:
    raise SystemExit(f"ERROR: ADMIN_ID must be a number, got: {_admin_id!r}")

PASSWORD_FILE = "fb_password.txt"
EXCEL_FILE    = "cookies.xlsx"

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)

WAITING_PW    = 1
WAITING_PHONE = 2

# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────

def get_saved_password():
    if os.path.exists(PASSWORD_FILE):
        with open(PASSWORD_FILE, "r") as f:
            return f.read().strip()
    return None

def save_password(pw):
    with open(PASSWORD_FILE, "w") as f:
        f.write(pw)

def save_cookies_to_excel(phone, cookies):
    cookies_str = json.dumps(cookies)

    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cookies"
        ws["A1"] = "Row"
        ws["B1"] = "Phone"
        ws["C1"] = "Cookies"

    next_row = ws.max_row + 1
    ws[f"A{next_row}"] = next_row - 1
    ws[f"B{next_row}"] = phone
    ws[f"C{next_row}"] = cookies_str
    wb.save(EXCEL_FILE)

def get_driver():
    # selenium/standalone-chrome image has chrome + chromedriver pre-installed
    # and on PATH. Let Selenium find them automatically via Service().
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    return webdriver.Chrome(options=options)

def do_facebook_login(phone, password):
    driver = get_driver()
    try:
        driver.get("https://www.facebook.com")
        wait = WebDriverWait(driver, 20)
        time.sleep(3)  # let page fully render

        # Dismiss any consent/cookie popups — try multiple times
        for _ in range(3):
            try:
                btn = driver.find_element(By.XPATH,
                    "//*[self::button or self::a or self::div]"
                    "[@role='button' or self::button]"
                    "[contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'ACCEPT')"
                    " or contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'ALLOW')"
                    " or contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'DECLINE')"
                    " or contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'REJECT')"
                    " or contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'CLOSE')"
                    " or contains(translate(text(),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'ONLY ALLOW')"
                    "]"
                )
                btn.click()
                time.sleep(1)
            except Exception:
                break

        # Try m.facebook.com as fallback if login form not found on desktop
        email_field = None
        for url in ["https://www.facebook.com", "https://m.facebook.com"]:
            try:
                if email_field is None and url != "https://www.facebook.com":
                    driver.get(url)
                    time.sleep(3)
                email_field = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "email"))
                )
                logging.info(f"Login form found on: {driver.current_url}")
                break
            except Exception:
                email_field = None

        if email_field is None:
            # Save page source for debugging
            logging.error(f"Page source snippet: {driver.page_source[:500]}")
            driver.quit()
            return None, "Could not find login form on facebook.com or m.facebook.com. Facebook may be blocking the headless browser."

        email_field.clear()
        email_field.send_keys(phone)

        pass_field = driver.find_element(By.ID, "pass")
        pass_field.clear()
        pass_field.send_keys(password)

        driver.find_element(By.NAME, "login").click()

        # Wait for page to change after login attempt
        time.sleep(6)

        current_url = driver.current_url
        page_source = driver.page_source

        logging.info(f"After login URL: {current_url}")

        # ── Failure cases ──────────────────────────────────────────────────

        # Wrong password — Facebook shows an error message on the login page
        if "login" in current_url or "login" in current_url.split("?")[0]:
            # Try to grab the actual error text Facebook shows
            error_text = "Wrong password or login blocked by Facebook."
            try:
                err_el = driver.find_element(By.XPATH,
                    "//*[contains(@data-testid,'royal_login_error') or contains(@id,'error_box') or contains(@class,'_9ay7')]"
                )
                error_text = err_el.text.strip() or error_text
            except Exception:
                pass
            driver.quit()
            return None, f"Login failed (still on login page). {error_text}"

        if "checkpoint" in current_url:
            driver.quit()
            return None, "Login blocked by Facebook checkpoint / security check. Manual action needed."

        if "two_step_verification" in current_url or "2fa" in current_url.lower():
            driver.quit()
            return None, "Two-factor authentication required. Disable 2FA and try again."

        if "recover" in current_url or "help" in current_url:
            driver.quit()
            return None, f"Facebook redirected to recovery/help page: {current_url}"

        # ── Success — we should be on home/feed ────────────────────────────
        cookies = driver.get_cookies()

        if not cookies:
            driver.quit()
            return None, "Login seemed to succeed but no cookies were returned."

        # Extra sanity check: look for a known logged-in cookie
        cookie_names = [c["name"] for c in cookies]
        logging.info(f"Cookies received: {cookie_names}")

        if "c_user" not in cookie_names and "xs" not in cookie_names:
            driver.quit()
            return None, (
                f"Login may have failed — expected session cookies not found.\n"
                f"Current URL: {current_url}\n"
                f"Cookies found: {cookie_names}"
            )

        driver.quit()
        return cookies, None

    except Exception as e:
        logging.exception("Unexpected error during Facebook login")
        try:
            driver.quit()
        except Exception:
            pass
        return None, f"Unexpected error: {str(e)}"

# ─────────────────────────────────────────────
#  /start
# ─────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return
    pw = get_saved_password()
    pw_status = "✅ Password saved" if pw else "❌ No password saved"
    await update.message.reply_text(
        f"👋 *Facebook Session Bot*\n\n"
        f"🔑 {pw_status}\n\n"
        f"Commands:\n"
        f"/setpw — Save Facebook password\n"
        f"/add — Login & save cookies to Excel\n"
        f"/dl — Download Excel file\n"
        f"/dlt — Delete/reset Excel file",
        parse_mode="Markdown"
    )

# ─────────────────────────────────────────────
#  /setpw
# ─────────────────────────────────────────────

async def cmd_setpw(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return
    await update.message.reply_text(
        "🔑 *Send your Facebook password now.*\n\n/abort to cancel.",
        parse_mode="Markdown"
    )
    return WAITING_PW

async def receive_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return ConversationHandler.END
    pw = update.message.text.strip()
    if not pw:
        await update.message.reply_text("⚠️ Empty, try again.")
        return WAITING_PW
    save_password(pw)
    try:
        await update.message.delete()
    except Exception:
        pass
    await update.message.reply_text("✅ *Password saved!*", parse_mode="Markdown")
    return ConversationHandler.END

# ─────────────────────────────────────────────
#  /add
# ─────────────────────────────────────────────

async def cmd_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return
    if not get_saved_password():
        await update.message.reply_text("⚠️ No password saved. Use /setpw first.")
        return ConversationHandler.END
    await update.message.reply_text(
        "📱 *Send the Facebook phone number.*\n\nInclude country code e.g. +1234567890\n\n/abort to cancel.",
        parse_mode="Markdown"
    )
    return WAITING_PHONE

async def receive_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return ConversationHandler.END
    phone = update.message.text.strip()
    if not phone:
        await update.message.reply_text("⚠️ Empty, try again.")
        return WAITING_PHONE

    pw  = get_saved_password()
    msg = await update.message.reply_text("⏳ *Logging in, please wait...*", parse_mode="Markdown")

    cookies, error = do_facebook_login(phone, pw)

    if error:
        await msg.edit_text(
            f"❌ *Login failed!*\n\n`{error}`",
            parse_mode="Markdown"
        )
        return ConversationHandler.END

    save_cookies_to_excel(phone, cookies)
    await msg.edit_text(
        f"✅ *Login successful!*\n\n"
        f"🍪 {len(cookies)} cookies for `{phone}` saved to Excel.\n"
        f"Use /dl to download.",
        parse_mode="Markdown"
    )
    return ConversationHandler.END

# ─────────────────────────────────────────────
#  /dl
# ─────────────────────────────────────────────

async def cmd_dl(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return
    if not os.path.exists(EXCEL_FILE):
        await update.message.reply_text("⚠️ No Excel file yet. Use /add first.")
        return
    with open(EXCEL_FILE, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename="cookies.xlsx",
            caption="🍪 *Cookies Excel file*",
            parse_mode="Markdown"
        )

# ─────────────────────────────────────────────
#  /dlt
# ─────────────────────────────────────────────

async def cmd_dlt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != ADMIN_ID:
        return
    if os.path.exists(EXCEL_FILE):
        os.remove(EXCEL_FILE)
        await update.message.reply_text("🗑 *Excel file deleted.*", parse_mode="Markdown")
    else:
        await update.message.reply_text("⚠️ No Excel file to delete.")

# ─────────────────────────────────────────────
#  /abort
# ─────────────────────────────────────────────

async def abort_conv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ Cancelled.")
    return ConversationHandler.END

# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def main():
    app = (
        Application.builder()
        .token(BOT_TOKEN)
        .connect_timeout(30)
        .read_timeout(30)
        .build()
    )

    setpw_conv = ConversationHandler(
        entry_points=[CommandHandler("setpw", cmd_setpw)],
        states={WAITING_PW: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_password)]},
        fallbacks=[CommandHandler("abort", abort_conv)],
    )

    add_conv = ConversationHandler(
        entry_points=[CommandHandler("add", cmd_add)],
        states={WAITING_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_phone)]},
        fallbacks=[CommandHandler("abort", abort_conv)],
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(setpw_conv)
    app.add_handler(add_conv)
    app.add_handler(CommandHandler("dl",  cmd_dl))
    app.add_handler(CommandHandler("dlt", cmd_dlt))

    logging.info("Bot started.")
    app.run_polling(
        drop_pending_updates=True,   # ignore old queued updates on startup
        allowed_updates=Update.ALL_TYPES,
    )

if __name__ == "__main__":
    main()
